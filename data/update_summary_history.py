#!/usr/bin/env python3
"""
Summary 시트 히스토리 관리 스크립트
3행의 데이터를 5행에 복사하고, 날짜 변경 시 히스토리로 저장
"""
from openpyxl import load_workbook
from datetime import datetime, date
import sys

def update_summary_history(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=False)
    ws = wb['Summary']

    # Admin 시트에서 현재 값들 읽기 (계산된 값)
    wb_data = load_workbook(xlsx_path, data_only=True)
    ws_admin = wb_data['Admin']

    # 현재 날짜 (NOW() 공식의 결과)
    current_datetime = datetime.now()
    current_date = current_datetime.date()

    # A5 기존 날짜 확인 (저장된 날짜값, 공식 아님)
    a5_value = ws.cell(5, 1).value
    if a5_value is None:
        existing_date = None
    elif isinstance(a5_value, datetime):
        existing_date = a5_value.date()
    elif isinstance(a5_value, date):
        existing_date = a5_value
    elif isinstance(a5_value, (int, float)):
        try:
            # 엑셀 날짜 시리얼 번호 변환
            existing_date = datetime.fromordinal(int(a5_value) + 693594).date()
        except:
            existing_date = None
    else:
        existing_date = None

    print(f"현재 날짜 (NOW): {current_date}")
    print(f"저장된 날짜 (A5): {existing_date}")

    # 날짜 비교
    if existing_date is None or current_date != existing_date:
        print(f"→ 날짜 변경 감지! 히스토리 행 삽입")

        # 5행 앞에 새로운 행 삽입
        ws.insert_rows(5)

        # 5행에 3행의 모든 값 복사 (공식이 아닌 계산값)
        # A5: 현재 날짜 (날짜만, 시간 제외)
        ws.cell(5, 1).value = current_date

        # B5-F5: 3행의 공식 복사 (Admin 시트 참조)
        for col in range(2, 7):
            source_formula = ws.cell(3, col).value
            ws.cell(5, col).value = source_formula

        # G5: 비워두기
        ws.cell(5, 7).value = None

        # H5-M5, O5-T5: 배분 비율 값 복사
        for col in list(range(8, 14)) + list(range(15, 21)):
            source_value = ws.cell(3, col).value
            ws.cell(5, col).value = source_value

        print(f"✓ 5행에 3행 데이터 복사 완료 (날짜: {current_date})")

    else:
        print(f"→ 동일 날짜. 5행 값만 업데이트")

        # A5는 그대로 유지 (날짜 동일)

        # B5-F5: 최신 Admin 공식 참조값으로 업데이트
        for col in range(2, 7):
            source_formula = ws.cell(3, col).value
            ws.cell(5, col).value = source_formula

        # H5-M5, O5-T5: 배분 비율 최신값으로 업데이트
        for col in list(range(8, 14)) + list(range(15, 21)):
            source_value = ws.cell(3, col).value
            ws.cell(5, col).value = source_value

        print(f"✓ 5행 값 업데이트 완료")

    wb.save(xlsx_path)
    print(f"\n파일 저장: {xlsx_path}")

if __name__ == '__main__':
    path = sys.argv[1] if len(sys.argv) > 1 else './watchlist.xlsx'
    update_summary_history(path)
