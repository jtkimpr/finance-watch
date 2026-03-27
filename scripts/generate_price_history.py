#!/usr/bin/env python3
"""
watchlist.xlsx의 Price 시트에서 각 종목의 현재가와 과거가를 추출하여
data/price-history.json으로 저장합니다.
"""

import json
import openpyxl
from datetime import datetime, timedelta
from pathlib import Path

def generate_price_history():
    """watchlist.xlsx에서 가격 이력 데이터 추출 및 JSON 생성"""

    # 파일 경로 설정
    xlsx_path = Path(__file__).parent.parent / "data" / "watchlist.xlsx"
    json_path = Path(__file__).parent.parent / "data" / "price-history.json"

    if not xlsx_path.exists():
        print(f"Error: {xlsx_path} not found")
        return False

    try:
        # watchlist.xlsx 읽기
        wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)

        # Info와 Price 시트 확인
        if "Info" not in wb.sheetnames or "Price" not in wb.sheetnames:
            print("Error: Info or Price sheet not found")
            return False

        info_ws = wb["Info"]
        price_ws = wb["Price"]

        # Info 시트에서 종목명, 티커, 카테고리 추출 (row 2부터)
        holdings = {}
        for row_idx, row in enumerate(info_ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:  # 종목명이 없으면 종료
                break
            name = row[0]
            ticker = row[1]
            category = row[3]

            holdings[ticker] = {
                "name": name,
                "ticker": ticker,
                "category": category,
                "prices": {}
            }

        # Price 시트 헤더 추출 (어느 컬럼이 어느 종목인지 파악)
        header_row = list(price_ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        ticker_to_col = {}
        for col_idx, cell in enumerate(header_row):
            if cell in holdings:
                ticker_to_col[cell] = col_idx

        # Price 시트에서 과거 가격 추출 (row 2부터, 최대 100행)
        # row 2 = 오늘, row 3 = 1일전, row 4 = 2일전, ...
        price_rows = list(price_ws.iter_rows(min_row=2, max_row=102, values_only=True))

        for day_offset, price_row in enumerate(price_rows[:100]):
            # 첫 번째 컬럼은 날짜
            date_val = price_row[0]
            if not date_val:
                break

            # 날짜 파싱
            if isinstance(date_val, datetime):
                price_date = date_val.date()
            else:
                try:
                    price_date = datetime.strptime(str(date_val), "%m/%d/%Y").date()
                except:
                    continue

            # 각 종목의 가격 저장
            for ticker, col_idx in ticker_to_col.items():
                price = price_row[col_idx]
                if price and isinstance(price, (int, float)):
                    if day_offset == 0:
                        day_label = "current"
                    elif day_offset == 1:
                        day_label = "day_1"
                    elif day_offset == 7:
                        day_label = "day_7"
                    elif day_offset == 30:
                        day_label = "day_30"
                    elif day_offset == 60:
                        day_label = "day_60"
                    else:
                        continue

                    holdings[ticker]["prices"][day_label] = float(price)

        # 과거 데이터가 없으면 이전 행에서 찾기
        for ticker in holdings:
            prices = holdings[ticker]["prices"]

            # day_1 데이터가 없으면 다음 행들 찾기
            if "day_1" not in prices:
                for day_offset in range(1, min(10, len(price_rows))):
                    price_row = price_rows[day_offset]
                    if ticker in ticker_to_col:
                        col_idx = ticker_to_col[ticker]
                        price = price_row[col_idx]
                        if price and isinstance(price, (int, float)):
                            prices["day_1"] = float(price)
                            break

            # day_7 데이터가 없으면 찾기
            if "day_7" not in prices:
                for day_offset in range(5, min(15, len(price_rows))):
                    price_row = price_rows[day_offset]
                    if ticker in ticker_to_col:
                        col_idx = ticker_to_col[ticker]
                        price = price_row[col_idx]
                        if price and isinstance(price, (int, float)):
                            prices["day_7"] = float(price)
                            break

            # day_30 데이터가 없으면 찾기
            if "day_30" not in prices:
                for day_offset in range(25, min(40, len(price_rows))):
                    price_row = price_rows[day_offset]
                    if ticker in ticker_to_col:
                        col_idx = ticker_to_col[ticker]
                        price = price_row[col_idx]
                        if price and isinstance(price, (int, float)):
                            prices["day_30"] = float(price)
                            break

            # day_60 데이터가 없으면 찾기
            if "day_60" not in prices:
                for day_offset in range(55, min(100, len(price_rows))):
                    price_row = price_rows[day_offset]
                    if ticker in ticker_to_col:
                        col_idx = ticker_to_col[ticker]
                        price = price_row[col_idx]
                        if price and isinstance(price, (int, float)):
                            prices["day_60"] = float(price)
                            break

        # 증감률 계산
        for ticker in holdings:
            holding = holdings[ticker]
            current = holding["prices"].get("current")

            if current:
                for day_label in ["day_1", "day_7", "day_30", "day_60"]:
                    past = holding["prices"].get(day_label)
                    if past:
                        change_pct = ((current - past) / past) * 100
                        holding[f"{day_label}_change"] = round(change_pct, 2)
                    else:
                        holding[f"{day_label}_change"] = None

        # JSON 저장 (current 가격만 유지, 나머지는 change로)
        output = {}
        for ticker, holding in holdings.items():
            output[ticker] = {
                "name": holding["name"],
                "ticker": holding["ticker"],
                "category": holding["category"],
                "current_price": holding["prices"].get("current"),
                "changes": {
                    "day_1": holding.get("day_1_change"),
                    "day_7": holding.get("day_7_change"),
                    "day_30": holding.get("day_30_change"),
                    "day_60": holding.get("day_60_change"),
                }
            }

        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(output, f, indent=2, ensure_ascii=False)

        print(f"✅ Price history saved to {json_path}")
        print(f"   Total holdings: {len(output)}")

        return True

    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = generate_price_history()
    exit(0 if success else 1)
