#!/usr/bin/env python3
"""
backfill_summary.py
Summary 시트에 과거 날짜별 데이터를 Price 시트 기준으로 역산해서 채움.
1회성 실행 스크립트.

사용법: python3 scripts/backfill_summary.py
"""

import datetime
import os
import re
import openpyxl
from openpyxl.utils import column_index_from_string

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(BASE_DIR, "..", "data", "watchlist.xlsx")

START_DATE = datetime.date(2026, 1, 1)
END_DATE   = datetime.date(2026, 3, 23)

SUMMARY_CATEGORIES = ["Crypto", "Cash", "Gold", "Kor Stock", "US Stock", "US Bonds"]

KRX_EXCHANGES = {"KRX", "KOSPI", "KOSDAQ"}
SPECIAL = {
    "KRW":  {"currency": "KRW", "exchange": "—", "ticker": "KRW"},
    "USD":  {"currency": "USD", "exchange": "—", "ticker": "USD"},
    "USDT": {"currency": "USD", "exchange": "—", "ticker": "USDT"},
}


# ---------------------------------------------------------------------------
# Price 시트 히스토리 로드
# ---------------------------------------------------------------------------

def parse_date(val):
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    if isinstance(val, str):
        for fmt in ("%m/%d/%Y", "%Y-%m-%d"):
            try:
                return datetime.datetime.strptime(val.strip(), fmt).date()
            except ValueError:
                pass
    return None


def build_price_history(wb):
    """Price 시트 전체를 읽어 {date: {col_idx: price}} 반환.
    None 값은 fill-forward로 채움 (직전 유효 가격 사용).
    """
    ws = wb["Price"]
    max_col = ws.max_column

    # 오래된 날짜부터 오름차순으로 읽기
    raw = []
    for row_idx in range(2, ws.max_row + 1):
        d = parse_date(ws.cell(row_idx, 1).value)
        if d is None:
            continue
        prices = {}
        for col_idx in range(2, max_col + 1):
            val = ws.cell(row_idx, col_idx).value
            if isinstance(val, (int, float)):
                prices[col_idx] = float(val)
        raw.append((d, prices))

    raw.sort(key=lambda x: x[0])

    # fill-forward: 이전 날짜의 유효 가격을 누적
    filled = {}
    last = {}
    for d, prices in raw:
        last.update(prices)
        filled[d] = dict(last)

    return filled


# ---------------------------------------------------------------------------
# 포트폴리오 계산 (generate_portfolio_json.py 로직 재사용)
# ---------------------------------------------------------------------------

def build_info_lookup(wb):
    ws = wb["Info"]
    lookup = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        name     = str(row[0]).strip()
        symbol   = str(row[1]).strip() if row[1] else ""
        exchange = str(row[2]).strip() if row[2] else ""
        currency = "KRW" if exchange in KRX_EXCHANGES else "USD"
        entry = {"currency": currency, "exchange": exchange, "ticker": symbol or name}
        lookup[name] = entry
        if symbol:
            lookup[symbol] = entry
    return lookup


def build_info_price_lookup(ws_info, price_lookup):
    """Info E열 수식에서 Price 컬럼 번호를 추출해 price_lookup으로 매핑."""
    info_prices = {}
    for row_idx in range(1, ws_info.max_row + 1):
        val = ws_info.cell(row_idx, 5).value
        if val is None:
            continue
        if isinstance(val, (int, float)):
            info_prices[row_idx] = float(val)
        elif isinstance(val, str):
            m = re.match(r"=INDEX\(Price!([A-Z]+):", val)
            if m:
                col_idx = column_index_from_string(m.group(1))
                info_prices[row_idx] = price_lookup.get(col_idx, 0.0)
    return info_prices


def resolve_N(row_num, ws_admin, info_prices, visited=None):
    if visited is None:
        visited = set()
    if row_num in visited:
        return 0.0
    visited.add(row_num)
    cell_val = ws_admin.cell(row_num, 14).value
    if cell_val is None:
        return 0.0
    if isinstance(cell_val, (int, float)):
        return float(cell_val)
    m = re.match(r"=Info!E(\d+)", str(cell_val))
    if m:
        return info_prices.get(int(m.group(1)), 0.0)
    m = re.match(r"=N(\d+)", str(cell_val))
    if m:
        return resolve_N(int(m.group(1)), ws_admin, info_prices, visited)
    return 0.0


def resolve_O(row_num, ws_admin, info_prices, visited=None):
    if visited is None:
        visited = set()
    if row_num in visited:
        return 1.0
    visited.add(row_num)
    val = ws_admin.cell(row_num, 15).value
    if val is None:
        return 1.0
    if isinstance(val, (int, float)):
        return float(val)
    m = re.match(r"=Info!E(\d+)", str(val))
    if m:
        return info_prices.get(int(m.group(1)), 1.0)
    m = re.match(r"=O(\d+)", str(val))
    if m:
        return resolve_O(int(m.group(1)), ws_admin, info_prices, visited)
    return 1.0


def compute_portfolios(ws_admin, ws_info, price_lookup):
    """특정 날짜의 price_lookup으로 멤버별 포트폴리오 계산."""
    info_prices = build_info_price_lookup(ws_info, price_lookup)

    portfolios = {}
    current_section = None

    for row_idx in range(1, ws_admin.max_row + 1):
        a = ws_admin.cell(row_idx, 1).value
        b = ws_admin.cell(row_idx, 2).value

        if a is None and b is None:
            continue
        if b is None and a is not None:
            a_str = str(a).strip()
            if a_str in ("KRW value",):
                continue
            current_section = a_str
            portfolios[current_section] = []
            continue
        if current_section is None or b is None:
            continue
        if str(b).strip() in ("Stock # ", "Stock price", "Fiat rate", "KRW value"):
            continue

        category = str(a).strip()

        def _parse_qty(val):
            if val is None:
                return 0.0
            if isinstance(val, (int, float)):
                return float(val)
            if isinstance(val, str) and re.match(r'^=[\d\s\+\-\*\/\.]+$', val):
                return float(eval(val[1:]))
            return 0.0

        qty = sum(_parse_qty(ws_admin.cell(row_idx, c).value) for c in range(6, 13))
        price = resolve_N(row_idx, ws_admin, info_prices)
        exchange_rate = resolve_O(row_idx, ws_admin, info_prices)
        portfolios[current_section].append({
            "category": category,
            "valuation": qty * price * exchange_rate,
        })

    return portfolios


def compute_summary_row(portfolios):
    """포트폴리오 dict → Summary 시트 한 행 데이터."""
    susie    = portfolios.get("Susie", [])
    dnb      = portfolios.get("Dirac & Broglie", [])
    jintae   = portfolios.get("Jintae", [])
    hyunhee  = portfolios.get("Hyunhee", [])

    susie_total   = sum(h["valuation"] for h in susie)
    dnb_total     = sum(h["valuation"] for h in dnb)
    jintae_total  = sum(h["valuation"] for h in jintae)
    hyunhee_total = sum(h["valuation"] for h in hyunhee)
    grand_total   = susie_total + dnb_total + jintae_total + hyunhee_total

    all_holdings = susie + dnb + jintae + hyunhee
    cat_totals = {}
    for h in all_holdings:
        cat_totals[h["category"]] = cat_totals.get(h["category"], 0) + h["valuation"]

    dnb_cat = {}
    for h in dnb:
        dnb_cat[h["category"]] = dnb_cat.get(h["category"], 0) + h["valuation"]

    return {
        "grand_total": round(grand_total),
        "susie":       round(susie_total),
        "dnb":         round(dnb_total),
        "jintae":      round(jintae_total),
        "hyunhee":     round(hyunhee_total),
        "all_cat":     {cat: round(cat_totals.get(cat, 0)) for cat in SUMMARY_CATEGORIES},
        "dnb_cat":     {cat: round(dnb_cat.get(cat, 0))    for cat in SUMMARY_CATEGORIES},
    }


# ---------------------------------------------------------------------------
# Summary 시트 쓰기
# ---------------------------------------------------------------------------

def write_summary_rows(ws, rows):
    """rows: [(date, row_dict), ...] 내림차순 정렬된 리스트.
    row 6 이하를 모두 지우고 새로 씀.
    """
    if ws.max_row >= 6:
        ws.delete_rows(6, ws.max_row - 5)

    NUM_FMT = '_-* #,##0_-;\\-* #,##0_-;_-* "-"_-;_-@_-'

    for d, row in rows:
        r = ws.max_row + 1
        ws.cell(r, 1).value = datetime.datetime(d.year, d.month, d.day)
        ws.cell(r, 1).number_format = 'yyyy\\-mm\\-dd'
        for col, val in [(2, row["grand_total"]), (3, row["susie"]), (4, row["dnb"]),
                         (5, row["jintae"]), (6, row["hyunhee"])]:
            ws.cell(r, col).value = val
            ws.cell(r, col).number_format = NUM_FMT
        for i, cat in enumerate(SUMMARY_CATEGORIES):
            ws.cell(r, 8  + i).value = row["all_cat"][cat]
            ws.cell(r, 8  + i).number_format = NUM_FMT
            ws.cell(r, 15 + i).value = row["dnb_cat"][cat]
            ws.cell(r, 15 + i).number_format = NUM_FMT


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("=== Summary 백필 시작 ===")
    print(f"대상 기간: {START_DATE} ~ {END_DATE}")

    print("\n[1/4] 파일 로드...")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=False)
    ws_admin = wb["Admin"]
    ws_info  = wb["Info"]

    print("[2/4] Price 시트 히스토리 읽기 (fill-forward)...")
    price_history = build_price_history(wb)
    target_dates = sorted(
        [d for d in price_history if START_DATE <= d <= END_DATE],
        reverse=True
    )
    print(f"  → Price 시트 총 {len(price_history)}개 날짜")
    print(f"  → 대상 날짜 {len(target_dates)}개: {target_dates[-1]} ~ {target_dates[0]}")

    print("\n[3/4] 날짜별 포트폴리오 계산...")
    rows = []
    for i, d in enumerate(target_dates):
        portfolios = compute_portfolios(ws_admin, ws_info, price_history[d])
        row = compute_summary_row(portfolios)
        rows.append((d, row))
        if (i + 1) % 20 == 0 or (i + 1) == len(target_dates):
            print(f"  {i+1}/{len(target_dates)} 완료  (예: {d} total={row['grand_total']:,})")

    print("\n[4/4] Summary 시트 쓰기...")
    ws_summary = wb["Summary"]
    write_summary_rows(ws_summary, rows)
    wb.save(XLSX_PATH)
    print(f"  → {len(rows)}개 행 저장 완료")

    print("\n[결과 미리보기]")
    for r in range(5, min(12, ws_summary.max_row + 1)):
        d_val    = ws_summary.cell(r, 1).value
        total    = ws_summary.cell(r, 2).value
        h_crypto = ws_summary.cell(r, 8).value or 0
        if total:
            print(f"  row{r:3d}: {str(d_val)[:10]}  total={total:>15,}  crypto={h_crypto:>15,}")
        else:
            print(f"  row{r:3d}: (공식 행)")

    print("\n=== 완료 ===")


if __name__ == "__main__":
    main()
