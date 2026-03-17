"""
update_watchlist.py
-------------------
Fetches latest market data and updates data/watchlist.xlsx (Price & Volume sheets).
Run daily via GitHub Actions or manually.
"""

import os
import time
import datetime
from pathlib import Path

import pytz
import requests
import yfinance as yf
from pykrx import stock as pykrx_stock
import finnhub
import openpyxl

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

KST = pytz.timezone("Asia/Seoul")

# Excel file path relative to this script (scripts/ -> .. -> data/)
SCRIPT_DIR = Path(__file__).parent
EXCEL_PATH = SCRIPT_DIR.parent / "data" / "watchlist.xlsx"

FINNHUB_API_KEY = os.environ["FINNHUB_API_KEY"]

# ---------------------------------------------------------------------------
# Column definitions
# Price sheet columns (in order):
#   Date, 0085P0, 182480, 232080, 360200, 361580, 379800, 379810, 411060,
#   438080, 459580, 475630, 476760, Bitcoin, Ethereum, QQQM, SCHD, SGOV,
#   SPYM, Strategy, Strategy Prf, Strategy Prf A, Strategy Prf D,
#   Strategy Prf F, UB Care, USDT_KRW Upbit, USD_KRW, VGIT, VGLT, VIG, VTIP
#
# Volume sheet columns (same but NO USD_KRW):
#   Date, 0085P0, ..., VTIP  (USD_KRW column absent)
# ---------------------------------------------------------------------------

PRICE_COLUMNS = [
    "Date",
    "0085P0", "182480", "232080", "360200", "361580",
    "379800", "379810", "411060", "438080", "459580",
    "475630", "476760",
    "Bitcoin", "Ethereum",
    "QQQM", "SCHD", "SGOV", "SPYM",
    "Strategy", "Strategy Prf", "Strategy Prf A", "Strategy Prf D",
    "Strategy Prf F",
    "UB Care",
    "USDT_KRW Upbit",
    "USD_KRW",
    "VGIT", "VGLT", "VIG", "VTIP",
]

VOLUME_COLUMNS = [
    "Date",
    "0085P0", "182480", "232080", "360200", "361580",
    "379800", "379810", "411060", "438080", "459580",
    "475630", "476760",
    "Bitcoin", "Ethereum",
    "QQQM", "SCHD", "SGOV", "SPYM",
    "Strategy", "Strategy Prf", "Strategy Prf A", "Strategy Prf D",
    "Strategy Prf F",
    "UB Care",
    "USDT_KRW Upbit",
    # USD_KRW intentionally absent
    "VGIT", "VGLT", "VIG", "VTIP",
]

# pykrx ticker codes (key = Excel column name)
PYKRX_TICKERS = {
    "0085P0": "0085P0",
    "182480": "182480",
    "232080": "232080",
    "360200": "360200",
    "361580": "361580",
    "379800": "379800",
    "379810": "379810",
    "411060": "411060",
    "438080": "438080",
    "459580": "459580",
    "475630": "475630",
    "476760": "476760",
    "UB Care": "032620",
}

# yfinance ticker codes
YFINANCE_TICKERS = {
    "Bitcoin":   "BTC-USD",
    "Ethereum":  "ETH-USD",
    "QQQM":      "QQQM",
    "SCHD":      "SCHD",
    "SGOV":      "SGOV",
    "SPYM":      "SPYM",
    "Strategy":  "MSTR",
    "USD_KRW":   "KRW=X",
    "VGIT":      "VGIT",
    "VGLT":      "VGLT",
    "VIG":       "VIG",
    "VTIP":      "VTIP",
}

# finnhub ticker codes
FINNHUB_TICKERS = {
    "Strategy Prf":   "STRK",
    "Strategy Prf A": "STRC",
    "Strategy Prf D": "STRD",
    "Strategy Prf F": "STRF",
}


# ---------------------------------------------------------------------------
# Data fetching
# ---------------------------------------------------------------------------

def fetch_pykrx(today_str: str) -> dict:
    """Fetch close price and volume for all pykrx tickers.
    today_str: YYYYMMDD format.
    Returns dict: {col_name: {"price": val, "volume": val}}
    """
    results = {}
    for col, code in PYKRX_TICKERS.items():
        try:
            df = pykrx_stock.get_market_ohlcv_by_date(today_str, today_str, code)
            if df is None or df.empty:
                print(f"  [pykrx] {col} ({code}): no data")
                results[col] = {"price": None, "volume": None}
            else:
                close = df["종가"].iloc[-1]
                volume = df["거래량"].iloc[-1]
                print(f"  [pykrx] {col} ({code}): close={close}, vol={volume}")
                results[col] = {"price": close, "volume": volume}
        except Exception as e:
            print(f"  [pykrx] {col} ({code}): ERROR — {e}")
            results[col] = {"price": None, "volume": None}
    return results


def fetch_yfinance() -> dict:
    """Fetch close price and volume for all yfinance tickers (last available candle).
    Returns dict: {col_name: {"price": val, "volume": val}}
    USD_KRW always has volume=None.
    """
    results = {}
    for col, ticker in YFINANCE_TICKERS.items():
        try:
            hist = yf.Ticker(ticker).history(period="5d")
            if hist is None or hist.empty:
                print(f"  [yfinance] {col} ({ticker}): no data")
                results[col] = {"price": None, "volume": None}
            else:
                row = hist.iloc[-1]
                price = row["Close"]
                volume = None if col == "USD_KRW" else row["Volume"]
                print(f"  [yfinance] {col} ({ticker}): close={price:.4f}, vol={volume}")
                results[col] = {"price": price, "volume": volume}
        except Exception as e:
            print(f"  [yfinance] {col} ({ticker}): ERROR — {e}")
            results[col] = {"price": None, "volume": None}
    return results


def fetch_finnhub() -> dict:
    """Fetch close price and volume for finnhub tickers (last daily candle, past 7 days).
    Returns dict: {col_name: {"price": val, "volume": val}}
    """
    client = finnhub.Client(api_key=FINNHUB_API_KEY)
    results = {}
    now_ts = int(time.time())
    from_ts = now_ts - 7 * 24 * 3600

    for col, ticker in FINNHUB_TICKERS.items():
        try:
            candles = client.stock_candles(ticker, "D", from_ts, now_ts)
            if candles is None or candles.get("s") != "ok" or not candles.get("c"):
                print(f"  [finnhub] {col} ({ticker}): no data")
                results[col] = {"price": None, "volume": None}
            else:
                price = candles["c"][-1]
                volume = candles["v"][-1]
                print(f"  [finnhub] {col} ({ticker}): close={price}, vol={volume}")
                results[col] = {"price": price, "volume": volume}
        except Exception as e:
            print(f"  [finnhub] {col} ({ticker}): ERROR — {e}")
            results[col] = {"price": None, "volume": None}
        time.sleep(0.2)

    return results


def fetch_coingecko_usdt_krw() -> dict:
    """Fetch USDT/KRW price from CoinGecko and Upbit volume.
    Returns {"price": val, "volume": val}
    """
    price = None
    volume = None

    # Price
    try:
        resp = requests.get(
            "https://api.coingecko.com/api/v3/simple/price",
            params={"ids": "tether", "vs_currencies": "krw"},
            timeout=10,
        )
        resp.raise_for_status()
        data = resp.json()
        price = data["tether"]["krw"]
        print(f"  [coingecko] USDT_KRW price: {price}")
    except Exception as e:
        print(f"  [coingecko] USDT_KRW price ERROR — {e}")

    # Volume (Upbit USDT/KRW)
    try:
        resp = requests.get(
            "https://api.coingecko.com/api/v3/exchanges/upbit/tickers",
            params={"coin_ids": "tether"},
            timeout=15,
        )
        resp.raise_for_status()
        tickers = resp.json().get("tickers", [])
        match = next(
            (t for t in tickers if t.get("base") == "USDT" and t.get("target") == "KRW"),
            None,
        )
        if match:
            volume = match.get("volume")
            print(f"  [coingecko] USDT_KRW Upbit volume: {volume}")
        else:
            print("  [coingecko] USDT_KRW Upbit volume: no USDT/KRW ticker found")
    except Exception as e:
        print(f"  [coingecko] USDT_KRW Upbit volume ERROR — {e}")

    return {"price": price, "volume": volume}


# ---------------------------------------------------------------------------
# Excel update
# ---------------------------------------------------------------------------

def date_matches(cell_value, today: datetime.date) -> bool:
    """Check whether the date stored in an Excel cell equals today.
    Handles both string 'MM/DD/YYYY' and datetime.date / datetime.datetime types.
    """
    if cell_value is None:
        return False
    if isinstance(cell_value, (datetime.datetime, datetime.date)):
        cell_date = cell_value if isinstance(cell_value, datetime.date) else cell_value.date()
        return cell_date == today
    if isinstance(cell_value, str):
        try:
            parsed = datetime.datetime.strptime(cell_value.strip(), "%m/%d/%Y").date()
            return parsed == today
        except ValueError:
            return False
    return False


def build_price_row(today_str_excel: str, pykrx_data: dict, yf_data: dict,
                    fh_data: dict, cg_data: dict) -> list:
    """Build a list of values for the Price sheet row in PRICE_COLUMNS order."""
    all_data = {}
    all_data.update({col: pykrx_data[col]["price"] for col in pykrx_data})
    all_data.update({col: yf_data[col]["price"] for col in yf_data})
    all_data.update({col: fh_data[col]["price"] for col in fh_data})
    all_data["USDT_KRW Upbit"] = cg_data["price"]

    row = []
    for col in PRICE_COLUMNS:
        if col == "Date":
            row.append(today_str_excel)
        else:
            row.append(all_data.get(col))
    return row


def build_volume_row(today_str_excel: str, pykrx_data: dict, yf_data: dict,
                     fh_data: dict, cg_data: dict) -> list:
    """Build a list of values for the Volume sheet row in VOLUME_COLUMNS order."""
    all_data = {}
    all_data.update({col: pykrx_data[col]["volume"] for col in pykrx_data})
    all_data.update({col: yf_data[col]["volume"] for col in yf_data})
    all_data.update({col: fh_data[col]["volume"] for col in fh_data})
    all_data["USDT_KRW Upbit"] = cg_data["volume"]

    row = []
    for col in VOLUME_COLUMNS:
        if col == "Date":
            row.append(today_str_excel)
        else:
            row.append(all_data.get(col))
    return row


def update_sheet(ws, row_data: list, today: datetime.date) -> None:
    """Insert or overwrite row 2 in the given worksheet.
    - If row 2 date == today: overwrite in place.
    - Otherwise: insert a new row at position 2.
    """
    current_date_cell = ws.cell(row=2, column=1).value
    if date_matches(current_date_cell, today):
        print(f"    Row 2 date matches today — overwriting in place.")
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=2, column=col_idx).value = value
    else:
        print(f"    Row 2 date is '{current_date_cell}' — inserting new row at position 2.")
        ws.insert_rows(2)
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=2, column=col_idx).value = value


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    now_kst = datetime.datetime.now(KST)
    today = now_kst.date()
    today_str_excel = today.strftime("%m/%d/%Y")   # MM/DD/YYYY for Excel
    today_str_pykrx = today.strftime("%Y%m%d")     # YYYYMMDD for pykrx

    print(f"=== Watchlist updater started ===")
    print(f"KST datetime : {now_kst.strftime('%Y-%m-%d %H:%M:%S %Z')}")
    print(f"Target date  : {today_str_excel}")
    print(f"Excel path   : {EXCEL_PATH}")
    print()

    # --- Fetch data ---
    print("[1/4] Fetching pykrx data...")
    pykrx_data = fetch_pykrx(today_str_pykrx)

    print("\n[2/4] Fetching yfinance data...")
    yf_data = fetch_yfinance()

    print("\n[3/4] Fetching finnhub data...")
    fh_data = fetch_finnhub()

    print("\n[4/4] Fetching CoinGecko / Upbit data...")
    cg_data = fetch_coingecko_usdt_krw()

    # --- Build rows ---
    price_row = build_price_row(today_str_excel, pykrx_data, yf_data, fh_data, cg_data)
    volume_row = build_volume_row(today_str_excel, pykrx_data, yf_data, fh_data, cg_data)

    # --- Update Excel ---
    print(f"\n[Excel] Opening {EXCEL_PATH} ...")
    wb = openpyxl.load_workbook(EXCEL_PATH)

    print("[Excel] Updating Price sheet...")
    ws_price = wb["Price"]
    update_sheet(ws_price, price_row, today)

    print("[Excel] Updating Volume sheet...")
    ws_volume = wb["Volume"]
    update_sheet(ws_volume, volume_row, today)

    wb.save(EXCEL_PATH)
    print(f"\n[Excel] Saved successfully to {EXCEL_PATH}")
    print(f"=== Done ===")


if __name__ == "__main__":
    main()
