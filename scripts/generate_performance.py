#!/usr/bin/env python3
"""
watchlist.xlsx의 Summary 시트에서 성과 데이터를 추출하여
data/performance.json으로 저장합니다.

출력 형식:
{
  "date": "YYYY-MM-DD",
  "members": {
    "Total":   { "current": 숫자, "changes": { "day_1": %, "day_7": %, "day_30": %, "day_60": % } },
    "D&B":     { ... },
    "Susie":   { ... },
    "Jintae":  { ... },
    "Hyunhee": { ... }
  }
}
"""

import json
import openpyxl
from datetime import datetime, timedelta
from pathlib import Path

# Summary 시트 컬럼 인덱스 → 멤버 키 매핑 (0-based)
# A=0: Date, B=1: Total, C=2: Susie, D=3: D&B, E=4: Jintae, F=5: Hyunhee
MEMBER_COLS = {
    "Total":   1,
    "Susie":   2,
    "D&B":     3,
    "Jintae":  4,
    "Hyunhee": 5,
}

def generate_performance():
    """watchlist.xlsx에서 멤버별 성과 데이터 추출 및 JSON 생성"""

    xlsx_path = Path(__file__).parent.parent / "data" / "watchlist.xlsx"
    json_path = Path(__file__).parent.parent / "data" / "performance.json"

    if not xlsx_path.exists():
        print(f"Error: {xlsx_path} not found")
        return False

    try:
        wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
        if "Summary" not in wb.sheetnames:
            print("Error: Summary sheet not found in watchlist.xlsx")
            return False

        ws = wb["Summary"]

        # 날짜별 멤버별 값 수집 (row 4부터)
        series = {}
        for row in ws.iter_rows(min_row=4, values_only=True):
            date_val = row[0]
            if not date_val or not isinstance(date_val, datetime):
                continue
            date = date_val.date()
            entry = {}
            for member, col_idx in MEMBER_COLS.items():
                val = row[col_idx]
                if isinstance(val, (int, float)):
                    entry[member] = float(val)
            if entry:
                series[date] = entry

        if not series:
            print("Error: No data found in Summary sheet")
            return False

        today = max(series.keys())
        past_dates = sorted([d for d in series if d < today], reverse=True)

        def find_closest(days_ago):
            target = today - timedelta(days=days_ago)
            best = min(past_dates, key=lambda d: abs((d - target).days), default=None)
            if best is None or abs((best - target).days) > 5:
                return None
            return best

        def calc_change(member, days_ago):
            current = series[today].get(member)
            if current is None:
                return None
            past_date = find_closest(days_ago)
            if past_date is None:
                return None
            past = series[past_date].get(member)
            if past is None or past == 0:
                return None
            return round((current - past) / past * 100, 2)

        members_out = {}
        for member in MEMBER_COLS:
            current = series[today].get(member)
            members_out[member] = {
                "current": int(current) if current is not None else None,
                "changes": {
                    "day_1":  calc_change(member, 1),
                    "day_7":  calc_change(member, 7),
                    "day_30": calc_change(member, 30),
                    "day_60": calc_change(member, 60),
                },
            }

        output = {
            "date": today.isoformat(),
            "members": members_out,
        }

        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(output, f, indent=2, ensure_ascii=False)

        print(f"✅ Performance data saved to {json_path}")
        print(f"   Date: {output['date']}")
        for m, v in members_out.items():
            c = v['current']
            print(f"   {m:10s}: ₩{c:>15,}  {v['changes']}")

        return True

    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    success = generate_performance()
    exit(0 if success else 1)
