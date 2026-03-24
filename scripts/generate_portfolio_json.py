#!/usr/bin/env python3
"""
generate_portfolio_json.py
Admin 시트의 모든 포트폴리오 섹션을 읽어 data/portfolio.json 생성

수식 캐시 의존 제거:
- Price 시트에서 직접 최신 가격 조회 (update_watchlist.py가 실제 숫자로 기록)
- Admin F:L 열에서 실제 보유수량 합산
- valuation = qty * price * exchange_rate 로 Python에서 직접 계산
"""

import datetime
import json
import os
import re
import openpyxl
import pytz
from openpyxl.utils import column_index_from_string

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(BASE_DIR, "..", "data", "watchlist.xlsx")
JSON_PATH = os.path.join(BASE_DIR, "..", "data", "portfolio.json")

KRX_EXCHANGES = {"KRX", "KOSPI", "KOSDAQ"}

SPECIAL = {
    "KRW":  {"currency": "KRW", "exchange": "—", "ticker": "KRW"},
    "USD":  {"currency": "USD", "exchange": "—", "ticker": "USD"},
    "USDT": {"currency": "USD", "exchange": "—", "ticker": "USDT"},
}


def build_info_lookup(wb):
    """Info 시트에서 name/symbol → (currency, exchange, ticker) 매핑"""
    ws = wb["Info"]
    lookup = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        name     = str(row[0]).strip()
        symbol   = str(row[1]).strip() if row[1] else ""
        exchange = str(row[2]).strip() if row[2] else ""
        currency = "KRW" if exchange in KRX_EXCHANGES else "USD"
        entry = {"currency": currency, "exchange": exchange, "ticker": symbol or name}
        lookup[name] = entry
        if symbol:
            lookup[symbol] = entry
    return lookup


def build_price_lookup(wb):
    """Price 시트 각 컬럼에서 최신 실제 숫자값 추출 (col_index → price)"""
    ws = wb["Price"]
    prices = {}
    max_col = ws.max_column
    for col_idx in range(2, max_col + 1):
        for row_idx in range(2, min(ws.max_row + 1, 1000)):
            val = ws.cell(row_idx, col_idx).value
            if val is not None and isinstance(val, (int, float)):
                prices[col_idx] = float(val)
                break
    return prices


def build_info_price_lookup(wb, price_lookup):
    """Info 시트 행 번호 → 가격 매핑 (Price 시트 참조 수식 해석)"""
    ws = wb["Info"]
    info_prices = {}
    for row_idx in range(1, ws.max_row + 1):
        val = ws.cell(row_idx, 5).value  # E열 = Price
        if val is None:
            continue
        if isinstance(val, (int, float)):
            info_prices[row_idx] = float(val)
        elif isinstance(val, str):
            m = re.match(r"=INDEX\(Price!([A-Z]+):", val)
            if m:
                col_idx = column_index_from_string(m.group(1))
                info_prices[row_idx] = price_lookup.get(col_idx, 0.0)
    return info_prices


def resolve_N(row_num, ws_admin, info_prices, visited=None):
    """Admin N열(가격) 해석: =Info!E{n}, =N{n} 체인, 또는 리터럴 숫자"""
    if visited is None:
        visited = set()
    if row_num in visited:
        return 0.0
    visited.add(row_num)

    cell_val = ws_admin.cell(row_num, 14).value  # N열 = 14번째
    if cell_val is None:
        return 0.0
    if isinstance(cell_val, (int, float)):
        return float(cell_val)
    m = re.match(r"=Info!E(\d+)", str(cell_val))
    if m:
        return info_prices.get(int(m.group(1)), 0.0)
    m = re.match(r"=N(\d+)", str(cell_val))
    if m:
        return resolve_N(int(m.group(1)), ws_admin, info_prices, visited)
    return 0.0


def resolve_O(row_num, ws_admin, info_prices, visited=None):
    """Admin O열(환율) 해석: =Info!E{n}, =O{n} 체인, 또는 리터럴 숫자"""
    if visited is None:
        visited = set()
    if row_num in visited:
        return 1.0
    visited.add(row_num)

    val = ws_admin.cell(row_num, 15).value  # O열 = 15번째
    if val is None:
        return 1.0
    if isinstance(val, (int, float)):
        return float(val)
    m = re.match(r"=Info!E(\d+)", str(val))
    if m:
        return info_prices.get(int(m.group(1)), 1.0)
    m = re.match(r"=O(\d+)", str(val))
    if m:
        return resolve_O(int(m.group(1)), ws_admin, info_prices, visited)
    return 1.0


def parse_admin_sheet(wb, info_lookup, info_prices):
    ws = wb["Admin"]
    portfolios = {}
    current_section = None

    for row_idx in range(1, ws.max_row + 1):
        a = ws.cell(row_idx, 1).value
        b = ws.cell(row_idx, 2).value

        if a is None and b is None:
            continue

        # 섹션 헤더: B열이 None이고 A열에 이름
        if b is None and a is not None:
            a_str = str(a).strip()
            # 헤더행 레이블(열 제목행) 제외
            if a_str in ("KRW value",):
                continue
            current_section = a_str
            portfolios[current_section] = []
            continue

        if current_section is None or b is None:
            continue

        # 헤더 레이블행 건너뜀 (B열이 'Stock # ' 같은 헤더인 경우)
        b_str = str(b).strip()
        if b_str in ("Stock # ", "Stock price", "Fiat rate", "KRW value"):
            continue

        name     = b_str
        category = str(a).strip()

        # qty: F:L열 합산 (col 6~12)
        def _parse_qty(val):
            if val is None:
                return 0.0
            if isinstance(val, (int, float)):
                return float(val)
            if isinstance(val, str) and re.match(r'^=[\d\s\+\-\*\/\.]+$', val):
                return float(eval(val[1:]))
            return 0.0

        qty = sum(_parse_qty(ws.cell(row_idx, c).value) for c in range(6, 13))

        # price: N열 (col 14)
        price = resolve_N(row_idx, ws, info_prices)

        # exchange_rate: O열 (col 15)
        exchange_rate = resolve_O(row_idx, ws, info_prices)

        valuation = qty * price * exchange_rate

        # 메타 정보
        if name in SPECIAL:
            meta = SPECIAL[name]
        elif name in info_lookup:
            meta = info_lookup[name]
        else:
            is_latin = all(c.isascii() for c in name.replace(" ", ""))
            meta = {
                "currency": "USD" if is_latin else "KRW",
                "exchange": "—",
                "ticker":   name,
            }

        portfolios[current_section].append({
            "name":      name,
            "ticker":    meta["ticker"],
            "exchange":  meta["exchange"],
            "category":  category,
            "currency":  meta["currency"],
            "qty":       qty,
            "price":     price,
            "valuation": valuation,
        })

    return portfolios


KST = pytz.timezone("Asia/Seoul")

SUMMARY_CATEGORIES = ["Crypto", "Cash", "Gold", "Kor Stock", "US Stock", "US Bonds"]
SUMMARY_MEMBERS = ["Susie", "Dirac & Broglie", "Jintae", "Hyunhee"]


def update_summary_history(portfolios, wb):
    """Summary 시트 5행부터 날짜별 히스토리를 내림차순으로 저장.
    - 오늘 날짜가 이미 5행에 있으면 덮어쓰기 (하루 4번 실행 대응)
    - 다른 날짜면 5행에 새 행 삽입 (기존 데이터 아래로 밀림)
    - A열: 날짜만 (시간 없음)
    - B-F열: hardcoded 숫자 (수식 아님)
    - H-M열: 전체 카테고리 비중 (/ grand total)
    - O-T열: Susie 카테고리 비중 (/ grand total)
    """
    ws = wb["Summary"]

    susie = portfolios.get("Susie", [])
    dnb = portfolios.get("Dirac & Broglie", [])
    jintae = portfolios.get("Jintae", [])
    hyunhee = portfolios.get("Hyunhee", [])

    susie_total = sum(h["valuation"] for h in susie)
    dnb_total = sum(h["valuation"] for h in dnb)
    jintae_total = sum(h["valuation"] for h in jintae)
    hyunhee_total = sum(h["valuation"] for h in hyunhee)
    grand_total = susie_total + dnb_total + jintae_total + hyunhee_total

    # 카테고리별 합계
    all_holdings = susie + dnb + jintae + hyunhee
    cat_totals = {}
    for h in all_holdings:
        cat_totals[h["category"]] = cat_totals.get(h["category"], 0) + h["valuation"]

    dnb_cat_totals = {}
    for h in dnb:
        dnb_cat_totals[h["category"]] = dnb_cat_totals.get(h["category"], 0) + h["valuation"]

    today = datetime.datetime.now(KST).date()

    # 4행 날짜 확인 (row1=헤더, row2=라이브, row3=히스토리헤더, row4~=히스토리데이터)
    a4 = ws.cell(4, 1).value
    if isinstance(a4, datetime.datetime):
        row4_date = a4.date()
    elif isinstance(a4, datetime.date):
        row4_date = a4
    else:
        row4_date = None

    if row4_date != today:
        ws.insert_rows(4)

    r = 4
    NUM_FMT = '_-* #,##0_-;\\-* #,##0_-;_-* "-"_-;_-@_-'

    # A: 날짜만 (시간 없이)
    ws.cell(r, 1).value = datetime.datetime(today.year, today.month, today.day)
    ws.cell(r, 1).number_format = 'yyyy\\-mm\\-dd'

    # B-F: 멤버별 총액
    for col, val in [(2, grand_total), (3, susie_total), (4, dnb_total),
                     (5, jintae_total), (6, hyunhee_total)]:
        ws.cell(r, col).value = round(val)
        ws.cell(r, col).number_format = NUM_FMT

    # H-M (col 8-13): 전체 카테고리 절대값 (KRW)
    for i, cat in enumerate(SUMMARY_CATEGORIES):
        ws.cell(r, 8 + i).value = round(cat_totals.get(cat, 0))
        ws.cell(r, 8 + i).number_format = NUM_FMT

    # O-T (col 15-20): D&B 카테고리 절대값 (KRW)
    for i, cat in enumerate(SUMMARY_CATEGORIES):
        ws.cell(r, 15 + i).value = round(dnb_cat_totals.get(cat, 0))
        ws.cell(r, 15 + i).number_format = NUM_FMT


def load_prev_totals():
    """기존 portfolio.json에서 각 멤버 총액 계산 → _prev_totals로 저장"""
    if not os.path.exists(JSON_PATH):
        return None
    try:
        with open(JSON_PATH, "r", encoding="utf-8") as f:
            existing = json.load(f)
        members = ["Susie", "Dirac & Broglie", "Jintae", "Hyunhee"]
        prev = {}
        for m in members:
            holdings = existing.get(m, [])
            prev[m] = sum(h.get("valuation", 0) for h in holdings)
        return prev
    except Exception as e:
        print(f"[warn] prev_totals 로드 실패: {e}")
        return None


def main():
    prev_totals = load_prev_totals()

    # data_only=False: 수식 문자열을 직접 파싱하기 위해
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=False)
    info_lookup   = build_info_lookup(wb)
    price_lookup  = build_price_lookup(wb)
    info_prices   = build_info_price_lookup(wb, price_lookup)
    portfolios    = parse_admin_sheet(wb, info_lookup, info_prices)

    portfolios["_prev_totals"] = prev_totals

    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(portfolios, f, ensure_ascii=False, indent=2)

    summary = {k: len(v) if isinstance(v, list) else v for k, v in portfolios.items()}
    print(f"portfolio.json 생성 완료: {summary}")

    # Summary 시트 히스토리 업데이트
    portfolios_clean = {k: v for k, v in portfolios.items() if k != "_prev_totals"}
    update_summary_history(portfolios_clean, wb)
    wb.save(XLSX_PATH)
    print("Summary 히스토리 업데이트 완료")


if __name__ == "__main__":
    main()
